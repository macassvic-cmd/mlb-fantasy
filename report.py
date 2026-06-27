"""
MLB Fantasy Report Generator
Builds an Excel workbook (output\\MLB_Projections_<date>.xlsx) and an HTML
dashboard (output\\dashboard.html) from the most recent pipeline data, then
opens both.

Usage:
  python report.py              # most recent data file
  python report.py 2026-06-11   # specific date
"""

import json
import os
import shutil
import subprocess
import sys
import webbrowser
from collections import defaultdict
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import projections as proj
from scrapers.market_lines import get_market_lines, compute_pp_ud_ratio, match_lines

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER = PatternFill("solid", fgColor="305496")
FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_BOLD   = Font(bold=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Derived fields
# ---------------------------------------------------------------------------

def confidence_score(p):
    """0-100 score reflecting how much data backs this projection."""
    score = 0
    r14 = p.get("rolling_14d") or {}
    sc = p.get("statcast") or {}
    fg = p.get("fg_stats") or {}
    wx = p.get("weather") or {}

    if p.get("lineup_confirmed"):
        score += 30

    g14 = r14.get("games", 0) or 0
    if g14 >= 10:
        score += 20
    elif g14 >= 5:
        score += 10

    if sc.get("xwoba_14d") is not None:
        score += 20

    if fg.get("woba") is not None:
        score += 10

    if wx.get("weather_available"):
        score += 10

    if p.get("days_rest") is not None:
        score += 10

    return min(score, 100)


def platoon_edge(p):
    """Yes/No/N/A based on the real wOBA-vs-wOBA-against matchup, not just
    handedness - falls back to N/A when either side's split sample is too
    small to trust (see fangraphs.match_platoon_matchup)."""
    pl = p.get("platoon") or {}
    adv = pl.get("advantage")
    if adv is None:
        return "N/A"
    return "Yes" if adv == "batter" else "No"


def weather_icon(row):
    if row.get("wx_indoor"):
        return "🏟️"
    precip = row.get("wx_precip")
    wind = row.get("wx_wind")
    temp = row.get("wx_temp")
    if precip is not None and precip >= 40:
        return "🌧️"
    if wind is not None and wind >= 12:
        return "💨"
    if temp is not None and temp >= 85:
        return "☀️"
    return "⛅"


def card_tier(ud_pts):
    if ud_pts >= 8:
        return "green"
    if ud_pts >= 5:
        return "yellow"
    return "red"


_PACIFIC = ZoneInfo("America/Los_Angeles")


def game_time_pt(game_date_utc):
    """Format an MLB API 'gameDate' UTC ISO timestamp as e.g. '7:05 PM PT'.
    Returns None if missing/unparseable so callers can skip it cleanly."""
    if not game_date_utc:
        return None
    try:
        dt_utc = datetime.fromisoformat(game_date_utc.replace("Z", "+00:00"))
        dt_pt = dt_utc.astimezone(_PACIFIC)
        return dt_pt.strftime("%-I:%M %p PT") if os.name != "nt" else dt_pt.strftime("%#I:%M %p PT")
    except (ValueError, TypeError):
        return None


def _is_early_game(game_date_utc):
    """Returns True if the game starts before 18:00 UTC (= 11:00 AM PT in
    summer), the window where getaway-day morning games cluster."""
    if not game_date_utc:
        return False
    try:
        dt = datetime.fromisoformat(game_date_utc.replace("Z", "+00:00"))
        return dt.hour < 18
    except (ValueError, TypeError):
        return False


def edge_label(edge):
    """Classify our projection vs. the posted UD line into OVER/UNDER/NEUTRAL."""
    if edge is None:
        return None
    if edge > 0.5:
        return "over"
    if edge < -0.5:
        return "under"
    return "neutral"


def build_card(row):
    return {
        "name":   row["name"],
        "team":   row["team"],
        "order":  row["order"] or "-",
        "ud":     fmt_value(row["ud_pts"], "1f"),
        "pp":     fmt_value(row["pp_pts"], "1f"),
        "xwoba":  fmt_value(row["xwoba"], "3f"),
        "barrel": fmt_value(row["barrel_pct"], "1f"),
        "era":    fmt_value(row["opp_era"], "2f"),
        "wxIcon": weather_icon(row),
        "wxText": row["weather"],
        "park":   fmt_value(row["park_hr"], "2f"),
        "platoon": row["platoon_edge"] == "Yes",
        "platoonMatchup": {
            "batterWoba":    row.get("platoon_batter_woba"),
            "batterLabel":   row.get("platoon_batter_label"),
            "pitcherWoba":   row.get("platoon_pitcher_woba"),
            "pitcherLabel":  row.get("platoon_pitcher_label"),
            "advantage":     row.get("platoon_advantage"),
        } if row.get("platoon_advantage") else None,
        "adjusted": row.get("adjusted", False),
        "anchored": row.get("market_anchored", False),
        "noLinePenalty": row.get("no_line_penalty", False),
        "getawayDayRisk": row.get("getaway_day_risk", False),
        "projectedLineup": row.get("lineup_status") == "projected",
        "tier":   card_tier(row["ud_pts"]),
        "edge":   row.get("edge"),
        "udLine": row.get("ud_line"),
        "edgeLabel": edge_label(row.get("edge")),
        "gameTimePt":  row.get("game_time_pt"),
        "gameDateUtc": row.get("game_date_utc"),
    }


def weather_str(p):
    wx = p.get("weather") or {}
    if wx.get("is_indoor"):
        return "Indoor"
    if not wx.get("weather_available"):
        return "N/A"
    temp = wx.get("temp_f")
    wind = wx.get("wind_speed_mph")
    wd = wx.get("wind_direction", "") or ""
    parts = []
    if temp is not None:
        parts.append(f"{temp:.0f}°F")
    if wind is not None:
        parts.append(f"{wind:.0f}mph {wd}".strip())
    return " ".join(parts) if parts else "N/A"


def build_row(p):
    sc = p.get("statcast") or {}
    fg = p.get("fg_stats") or {}
    mt = p.get("matchup") or {}
    pf = p.get("park_factor") or {}
    wx = p.get("weather") or {}
    r7 = p.get("rolling_7d") or {}
    r14 = p.get("rolling_14d") or {}

    return {
        "player_id":    p.get("player_id"),
        "name":         p.get("name", ""),
        "team":         p.get("team_name", ""),
        "opp_team":     p.get("opp_team_name", ""),
        "pos":          p.get("position", ""),
        "order":        p.get("batting_order") or 0,
        "ud_pts":       proj.ud_fpts(p),
        "pp_pts":       proj.pp_fpts(p),
        "confidence":   confidence_score(p),
        "xwoba":        sc.get("xwoba_14d") if sc.get("xwoba_14d") is not None else sc.get("xwoba_30d"),
        "barrel_pct":   sc.get("barrel_pct_14d"),
        "hard_hit_pct": sc.get("hard_hit_pct_14d"),
        "ev":           sc.get("avg_ev_14d"),
        "r7":           r7.get("ud_fpts_per_game"),
        "r14":          r14.get("ud_fpts_per_game"),
        "opp_sp":       mt.get("pitcher_name", ""),
        "opp_era":      mt.get("era") if mt.get("era") is not None else mt.get("era_fg"),
        "opp_fip":      mt.get("fip"),
        "weather":      weather_str(p),
        "wx_indoor":    bool(wx.get("is_indoor")),
        "wx_temp":      wx.get("temp_f"),
        "wx_wind":      wx.get("wind_speed_mph"),
        "wx_dir":       wx.get("wind_direction"),
        "wx_precip":    wx.get("precip_probability"),
        "park_hr":      pf.get("hr"),
        "days_rest":    p.get("days_rest"),
        "platoon_edge": platoon_edge(p),
        "platoon_batter_woba":    (p.get("platoon") or {}).get("batter_woba"),
        "platoon_batter_label":   (p.get("platoon") or {}).get("batter_split_label"),
        "platoon_pitcher_woba":   (p.get("platoon") or {}).get("pitcher_woba_against"),
        "platoon_pitcher_label":  (p.get("platoon") or {}).get("pitcher_split_label"),
        "platoon_advantage":      (p.get("platoon") or {}).get("advantage"),
        "comp":         proj.composite_score(p, "ud"),
        "game_pk":      p.get("game_pk"),
        "home_away":    p.get("home_away"),
        "venue":        p.get("venue_name", ""),
        "lineup_status":    p.get("lineup_status", "confirmed"),
        "lineup_confirmed": bool(p.get("lineup_confirmed", True)),
        "game_date_utc":    p.get("game_date_utc"),
        "game_time_pt":     game_time_pt(p.get("game_date_utc")),
    }


def tier(value, thresholds):
    """thresholds = (green_min, yellow_min)"""
    green_min, yellow_min = thresholds
    if value >= green_min:
        return "green"
    if value >= yellow_min:
        return "yellow"
    return "red"


def percentile(values, pct):
    s = sorted(values)
    if not s:
        return 0
    idx = int(round((pct / 100.0) * (len(s) - 1)))
    return s[idx]


# ---------------------------------------------------------------------------
# Projection recalibration
#
# pipeline.py's raw "projected" stat lines run hot (it scales whole rolling
# windows rather than realistic per-game rates), which inflates UD/PP points
# well past real sportsbook lines (which sit under ~10-12 even for elite
# hitters). Rather than re-deriving the model, we map each player's raw UD
# points to a realistic target band based on their percentile rank for the
# day, and carry PP points along at the same raw PP/UD ratio.
# ---------------------------------------------------------------------------

# (percentile, target UD pts) anchor points -> tuned to match the real
# UD/PP "Fantasy Points" line distribution (roughly 3.5-9.5 for a typical
# slate), so raw projections land close to market lines before anchoring.
PTS_CURVE = [
    (0.05, 3.5),
    (0.25, 4.5),
    (0.50, 6.0),
    (0.75, 7.5),
    (0.95, 9.5),
]


def pts_target(pctile):
    for (p0, v0), (p1, v1) in zip(PTS_CURVE, PTS_CURVE[1:]):
        if pctile <= p1:
            frac = (pctile - p0) / (p1 - p0) if p1 > p0 else 0
            return v0 + frac * (v1 - v0)
    return PTS_CURVE[-1][1]


def recalibrate_points(rows):
    """Rescale row['ud_pts'] / row['pp_pts'] in place to realistic ranges."""
    n = len(rows)
    order = sorted(range(n), key=lambda i: rows[i]["ud_pts"])
    for rank, i in enumerate(order):
        pctile = rank / (n - 1) if n > 1 else 1.0
        target_ud = pts_target(pctile)
        raw_ud = rows[i]["ud_pts"]
        raw_pp = rows[i]["pp_pts"]
        ratio = raw_pp / raw_ud if raw_ud > 0.01 else 0.93
        rows[i]["ud_pts"] = round(target_ud, 2)
        rows[i]["pp_pts"] = round(target_ud * ratio, 2)


# ---------------------------------------------------------------------------
# Market line anchoring
#
# Underdog and PrizePicks post their own "Fantasy Points" lines for today's
# games. Those lines are the real benchmark - our projection should land
# close to them, not float off on its own percentile-based scale. For any
# player with a posted line, we keep our raw recalibrated value only as an
# "edge" (clamped) on top of that line. The clamp must stay above the
# Value Plays threshold (1.5 pts, see VALUE_PLAY_EDGE below) or no edge can
# ever qualify as a Value Play.
# ---------------------------------------------------------------------------

MARKET_EDGE_CLAMP = 2.0
VALUE_PLAY_EDGE = 1.5  # must stay below MARKET_EDGE_CLAMP or no edge can ever qualify
assert VALUE_PLAY_EDGE < MARKET_EDGE_CLAMP, "Value Plays threshold must be below the market edge clamp"

# Players without a posted UD/PP line are systematically more volatile than
# the market suggests — UD withholds lines when lineup status is uncertain.
# Discount their projection to reflect this hidden risk and make it harder
# for them to crack the Top 25 on model strength alone.
NO_LINE_PENALTY = 0.20

# Applied on top of NO_LINE_PENALTY when BOTH lineup_confirmed=False AND the
# game is an early start (before 18:00 UTC / 11 AM PT). These two signals
# compound each other: the market's silence + an unconfirmed lineup on a
# getaway-day morning game is the highest-risk profile we've identified.
# Combined total: ×0.80 × ×0.75 = ×0.60 (40% reduction from base).
GETAWAY_DAY_PENALTY = 0.25


def apply_market_anchor(rows, market_lines, market_corrections=None):
    """Anchor row['ud_pts'] / row['pp_pts'] to today's posted UD/PP lines
    where available. If a player has a market_correction factor (how much
    they've historically beaten/missed their own line), that factor is
    applied to the line itself before adding the signal-based edge. Returns
    the set of player_ids that were anchored (these should be skipped by
    apply_corrections to avoid double adjustment)."""
    market_corrections = market_corrections or {}
    anchored = set()
    pp_ud_ratio = compute_pp_ud_ratio(market_lines)

    for row in rows:
        ud_line, pp_line = match_lines(row["name"], market_lines, pp_ud_ratio)
        if ud_line is None and pp_line is None:
            row["market_anchored"] = False
            continue

        mc = market_corrections.get(str(row.get("player_id")))
        if mc:
            ud_line = ud_line * mc["ud"]
            pp_line = pp_line * mc["pp"]

        raw_ud, raw_pp = row["ud_pts"], row["pp_pts"]

        ud_edge = max(-MARKET_EDGE_CLAMP, min(MARKET_EDGE_CLAMP, raw_ud - ud_line))
        row["ud_pts"] = round(ud_line + ud_edge, 2)
        row["ud_line"] = round(ud_line, 2)
        row["edge"] = round(ud_edge, 2)

        pp_edge = max(-MARKET_EDGE_CLAMP, min(MARKET_EDGE_CLAMP, raw_pp - pp_line))
        row["pp_pts"] = round(pp_line + pp_edge, 2)
        row["pp_line"] = round(pp_line, 2)

        row["market_anchored"] = True
        anchored.add(str(row.get("player_id")))

    return anchored


VALUE_PLAYS_DIR = os.path.join("data", "value_plays")


def save_value_plays(date_str, value_rows):
    """Persist today's Value Plays calls so tracker.py can grade OVER/UNDER
    accuracy once actual results are in."""
    plays = []
    for row in value_rows:
        plays.append({
            "player_id": row.get("player_id"),
            "name": row["name"],
            "team": row["team"],
            "call": "over" if row["edge"] > 0 else "under",
            "edge": row["edge"],
            "ud_line": row.get("ud_line"),
        })

    os.makedirs(VALUE_PLAYS_DIR, exist_ok=True)
    out_path = os.path.join(VALUE_PLAYS_DIR, f"{date_str}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"date": date_str, "plays": plays}, f, indent=2)


# ---------------------------------------------------------------------------
# Per-player model correction
#
# tracker.py records projected vs. actual UD/PP points for every graded
# player-day in data/results/all_results.json. Once a player has enough
# graded games, we know whether our projection has been running hot or cold
# for them specifically, and nudge future projections by that ratio.
# ---------------------------------------------------------------------------

MIN_GRADED_GAMES = 10
CORRECTION_CLAMP = (0.8, 1.2)  # avoid wild swings from small/noisy samples


def build_corrections(results_data, min_games=MIN_GRADED_GAMES):
    """Return {player_id_str: {"ud": factor, "pp": factor}} for players with
    enough graded history (avg actual / avg projected, clamped).

    Players without enough live-tracked history yet fall back to season-long
    correction factors from the backtest (results_data["backtest_corrections"])."""
    corrections = {}
    lo, hi = CORRECTION_CLAMP
    results_data = results_data or {}

    for pid, entry in (results_data.get("players", {}) or {}).items():
        history = entry.get("history", [])
        if len(history) < min_games:
            continue

        ud_proj = sum(h["projected_ud"] for h in history)
        ud_actual = sum(h["actual_ud"] for h in history)
        pp_proj = sum(h["projected_pp"] for h in history)
        pp_actual = sum(h["actual_pp"] for h in history)

        ud_factor = ud_actual / ud_proj if ud_proj > 0 else 1.0
        pp_factor = pp_actual / pp_proj if pp_proj > 0 else 1.0

        corrections[str(pid)] = {
            "ud": round(min(max(ud_factor, lo), hi), 3),
            "pp": round(min(max(pp_factor, lo), hi), 3),
        }

    for pid, c in (results_data.get("backtest_corrections", {}) or {}).items():
        if pid not in corrections:
            corrections[pid] = {
                "ud": round(min(max(c["ud"], lo), hi), 3),
                "pp": round(min(max(c["pp"], lo), hi), 3),
            }

    return corrections


def apply_corrections(rows, corrections, skip=None):
    """Apply each player's personal correction factor to ud_pts/pp_pts in
    place and flag row['adjusted'] so the dashboard can badge it. Players in
    `skip` (already anchored to a posted market line) are left untouched."""
    skip = skip or set()
    for row in rows:
        if str(row.get("player_id")) in skip:
            row["adjusted"] = False
            continue
        c = corrections.get(str(row.get("player_id")))
        if not c:
            row["adjusted"] = False
            continue
        row["ud_pts"] = round(row["ud_pts"] * c["ud"], 2)
        row["pp_pts"] = round(row["pp_pts"] * c["pp"], 2)
        row["adjusted"] = True


def apply_no_line_penalty(rows, anchored_ids=None):
    """Apply a flat projection discount to every player without a posted
    UD/PP market line. The absence of a line is itself a signal — UD
    withholds lines when a player's lineup status is uncertain — so we
    reduce their projection by NO_LINE_PENALTY to make it harder for them
    to crowd out market-confirmed players in the Top 25.

    An additional GETAWAY_DAY_PENALTY is layered on top when the player
    also has lineup_confirmed=False AND an early game start (before 18:00
    UTC / 11 AM PT). These two signals compounding each other represents
    the highest-risk profile — market silence + unconfirmed lineup on a
    morning getaway game.

    Sets row['no_line_penalty'] and row['getaway_day_risk'] flags so the
    dashboard can badge them appropriately."""
    anchored_ids = anchored_ids or set()
    for row in rows:
        if str(row.get("player_id")) in anchored_ids or row.get("market_anchored"):
            row["no_line_penalty"] = False
            row["getaway_day_risk"] = False
            continue

        row["ud_pts"] = round(row["ud_pts"] * (1 - NO_LINE_PENALTY), 2)
        row["pp_pts"] = round(row["pp_pts"] * (1 - NO_LINE_PENALTY), 2)
        row["no_line_penalty"] = True

        unconfirmed = not row.get("lineup_confirmed", True)
        early = _is_early_game(row.get("game_date_utc"))
        if unconfirmed and early:
            row["ud_pts"] = round(row["ud_pts"] * (1 - GETAWAY_DAY_PENALTY), 2)
            row["pp_pts"] = round(row["pp_pts"] * (1 - GETAWAY_DAY_PENALTY), 2)
            row["getaway_day_risk"] = True
        else:
            row["getaway_day_risk"] = False


# ---------------------------------------------------------------------------
# Excel: Tab 1 + Tab 2 - leaderboards
# ---------------------------------------------------------------------------

LEADER_COLS = [
    ("name",         "Player",       28, "s"),
    ("team",         "Team",         22, "s"),
    ("order",        "Bat Order",    9,  "i"),
    ("ud_pts",       "Proj UD Pts",  11, "1f"),
    ("pp_pts",       "Proj PP Pts",  11, "1f"),
    ("confidence",   "Confidence",   10, "i"),
    ("xwoba",        "xwOBA",        8,  "3f"),
    ("barrel_pct",   "Barrel%",      8,  "1f"),
    ("hard_hit_pct", "HardHit%",     9,  "1f"),
    ("ev",           "EV",           7,  "1f"),
    ("r7",           "7d Fpts",      8,  "1f"),
    ("r14",          "14d Fpts",     8,  "1f"),
    ("opp_sp",       "Opp SP",       20, "s"),
    ("opp_era",      "Opp SP ERA",   10, "2f"),
    ("opp_fip",      "Opp SP FIP",   10, "2f"),
    ("weather",      "Weather",      16, "s"),
    ("park_hr",      "Park HR Fctr", 11, "2f"),
    ("days_rest",    "Days Rest",    9,  "i"),
    ("platoon_edge", "Platoon Edge", 12, "s"),
]


def fmt_value(val, kind):
    if val is None or val == "":
        return "N/A"
    if kind == "i":
        return int(val)
    if kind == "1f":
        return round(float(val), 1)
    if kind == "2f":
        return round(float(val), 2)
    if kind == "3f":
        return round(float(val), 3)
    return val


def write_leaderboard_sheet(wb, title, rows, ud_thresholds, color_rows=True, autofilter=False):
    ws = wb.create_sheet(title)

    for c, (_, header, width, _) in enumerate(LEADER_COLS, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, 2):
        t = tier(row["ud_pts"], ud_thresholds)
        fill = {"green": FILL_GREEN, "yellow": FILL_YELLOW, "red": FILL_RED}[t]
        for c, (key, _, _, kind) in enumerate(LEADER_COLS, 1):
            cell = ws.cell(row=r, column=c, value=fmt_value(row[key], kind))
            cell.border = BORDER
            if color_rows:
                cell.fill = fill

    if autofilter:
        last_col = get_column_letter(len(LEADER_COLS))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    return ws


# ---------------------------------------------------------------------------
# Excel: Tab 3 - Matchup Report
# ---------------------------------------------------------------------------

def write_matchup_sheet(wb, players_by_game):
    ws = wb.create_sheet("Matchup Report")
    ws.column_dimensions["A"].width = 4
    for col, w in zip("BCDE", (24, 9, 11, 11)):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["F"].width = 3
    for col, w in zip("GHIJ", (24, 9, 11, 11)):
        ws.column_dimensions[col].width = w

    r = 1
    for game_pk, players in players_by_game.items():
        away = [p for p in players if p["home_away"] == "away"]
        home = [p for p in players if p["home_away"] == "home"]
        away.sort(key=lambda p: p["order"] or 99)
        home.sort(key=lambda p: p["order"] or 99)

        away_team = away[0]["team"] if away else (home[0]["opp_team"] if home else "Away")
        home_team = home[0]["team"] if home else (away[0]["opp_team"] if away else "Home")
        venue = home[0]["venue"] if home else (away[0]["venue"] if away else "")
        wx = home[0]["weather"] if home else (away[0]["weather"] if away else "")

        title_cell = ws.cell(row=r, column=1, value=f"{away_team} @ {home_team}  |  {venue}  |  {wx}")
        title_cell.font = FONT_BOLD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        r += 1

        for col, header in zip("BCDE", ("Away Lineup", "Order", "UD Pts", "PP Pts")):
            cell = ws[f"{col}{r}"]
            cell.value = header
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = CENTER
        for col, header in zip("GHIJ", ("Home Lineup", "Order", "UD Pts", "PP Pts")):
            cell = ws[f"{col}{r}"]
            cell.value = header
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = CENTER
        r += 1

        n = max(len(away), len(home))
        for i in range(n):
            if i < len(away):
                p = away[i]
                ws.cell(row=r, column=2, value=p["name"])
                ws.cell(row=r, column=3, value=p["order"] or "").alignment = CENTER
                ws.cell(row=r, column=4, value=round(p["ud_pts"], 1))
                ws.cell(row=r, column=5, value=round(p["pp_pts"], 1))
            if i < len(home):
                p = home[i]
                ws.cell(row=r, column=7, value=p["name"])
                ws.cell(row=r, column=8, value=p["order"] or "").alignment = CENTER
                ws.cell(row=r, column=9, value=round(p["ud_pts"], 1))
                ws.cell(row=r, column=10, value=round(p["pp_pts"], 1))
            r += 1

        r += 2  # blank rows between games

    return ws


# ---------------------------------------------------------------------------
# Excel: Tab 4 - SP Report
# ---------------------------------------------------------------------------

SP_COLS = [
    ("name",   "Pitcher",       24, "s"),
    ("team",   "Faces Team",    22, "s"),
    ("era",    "ERA",           8,  "2f"),
    ("fip",    "FIP",           8,  "2f"),
    ("whip",   "WHIP",          8,  "2f"),
    ("k_pct",  "K%",            8,  "1pct"),
    ("hh_pct", "HardHit% Allowed", 16, "s"),
    ("hittable", "Hittable Score", 14, "1f"),
]


def write_sp_sheet(wb, raw_players):
    ws = wb.create_sheet("SP Report")

    pitchers = {}
    for p in raw_players:
        mt = p.get("matchup") or {}
        name = mt.get("pitcher_name")
        if not name or name in pitchers:
            continue
        whip = mt.get("whip") if mt.get("whip") is not None else mt.get("whip_fg")
        pitchers[name] = {
            "name": name,
            "team": p.get("opp_team_name", ""),
            "era": mt.get("era") if mt.get("era") is not None else mt.get("era_fg"),
            "fip": mt.get("fip"),
            "whip": whip,
            "k_pct": mt.get("k_pct"),
            "hh_pct": "N/A",
        }

    rows = list(pitchers.values())
    for row in rows:
        era = row["era"] if row["era"] is not None else 4.20
        fip = row["fip"] if row["fip"] is not None else 4.20
        whip = row["whip"] if row["whip"] is not None else 1.30
        k_pct = row["k_pct"] if row["k_pct"] is not None else 0.23
        row["hittable"] = round(float(era) + float(fip) + float(whip) * 3 - float(k_pct) * 15, 2)

    rows.sort(key=lambda r: r["hittable"], reverse=True)

    for c, (_, header, width, _) in enumerate(SP_COLS, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, 2):
        for c, (key, _, _, kind) in enumerate(SP_COLS, 1):
            val = row.get(key)
            if kind == "1pct" and val is not None:
                val = round(float(val) * 100, 1)
            elif kind in ("1f", "2f") and val is not None:
                val = round(float(val), 2 if kind == "2f" else 1)
            elif val is None:
                val = "N/A"
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER

    last_col = get_column_letter(len(SP_COLS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    return ws, pitchers


# ---------------------------------------------------------------------------
# Excel: Tab 5 - Stack Targets
# ---------------------------------------------------------------------------

STACK_COLS = [
    ("team",     "Team",        22, "s"),
    ("size",     "Stack Size",  10, "i"),
    ("players",  "Players (Bat Order)", 45, "s"),
    ("combined", "Combined UD Pts", 14, "1f"),
    ("opp_sp",   "Opp SP",      20, "s"),
]


def write_stack_sheet(wb, players_by_team):
    ws = wb.create_sheet("Stack Targets")

    stacks = []
    for team, players in players_by_team.items():
        ordered = sorted(
            (p for p in players if p["order"] and 1 <= p["order"] <= 9),
            key=lambda p: p["order"],
        )
        by_order = {p["order"]: p for p in ordered}

        for size in (2, 3):
            for start in range(1, 10):
                slots = [((start - 1 + i) % 9) + 1 for i in range(size)]  # wraps 9 -> 1
                if not all(s in by_order for s in slots):
                    continue
                grp = [by_order[s] for s in slots]
                combined = sum(p["ud_pts"] for p in grp)
                stacks.append({
                    "team": team,
                    "size": size,
                    "players": ", ".join(f"{p['name']} ({p['order']})" for p in grp),
                    "combined": combined,
                    "opp_sp": grp[0]["opp_sp"],
                })

    stacks.sort(key=lambda s: s["combined"], reverse=True)
    top_stacks = stacks[:15]

    for c, (_, header, width, _) in enumerate(STACK_COLS, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"

    for r, row in enumerate(top_stacks, 2):
        for c, (key, _, _, kind) in enumerate(STACK_COLS, 1):
            val = row[key]
            if kind == "1f":
                val = round(val, 1)
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER

    return ws


# ---------------------------------------------------------------------------
# HTML dashboard
# ---------------------------------------------------------------------------

DASHBOARD_COLS = [
    ("rank",        "Rank",         "i"),
    ("name",        "Player",       "s"),
    ("team",        "Team",         "s"),
    ("game_time_pt","Game Time",    "s"),
    ("order",       "Order",        "i"),
    ("ud_pts",      "UD Pts",       "1f"),
    ("pp_pts",      "PP Pts",       "1f"),
    ("xwoba",       "xwOBA",        "3f"),
    ("barrel_pct",  "Barrel%",      "1f"),
    ("hard_hit_pct","Hard Hit%",    "1f"),
    ("opp_era",     "Opp ERA",      "2f"),
    ("opp_fip",     "Opp FIP",      "2f"),
    ("weather",     "Weather",      "s"),
    ("park_hr",     "Park Factor",  "2f"),
    ("days_rest",   "Days Rest",    "i"),
    ("platoon_edge","Platoon Edge", "s"),
]


def write_dashboard(rows, date_str, out_path, results_data=None, top25_data=None):
    games_count = len({r["game_pk"] for r in rows})
    generated_dt = datetime.now()
    last_updated = generated_dt.strftime("%Y-%m-%d %I:%M %p")
    generated_at_iso = generated_dt.isoformat()
    player_count = len(rows)

    results_data = results_data or {"dates": {}, "players": {}}
    top25_data = top25_data or {"dates": {}, "players": {}}

    # --- Results: last 30 days calendar heatmap ------------------------
    result_dates = sorted(results_data.get("dates", {}).keys())[-30:]
    calendar_cells = []
    total_ud_hit = total_ud = total_pp_hit = total_pp = 0
    for d in result_dates:
        s = results_data["dates"][d]
        calendar_cells.append({
            "date": d,
            "ud_hit_rate": s["ud"]["hit_rate"],
            "pp_hit_rate": s["pp"]["hit_rate"],
            "player_count": s["player_count"],
        })
        total_ud_hit += s["ud"]["win"]
        total_ud += s["ud"]["total"]
        total_pp_hit += s["pp"]["win"]
        total_pp += s["pp"]["total"]
    calendar_js = json.dumps(calendar_cells)
    overall_ud_rate = round(100 * total_ud_hit / total_ud, 1) if total_ud else 0.0
    overall_pp_rate = round(100 * total_pp_hit / total_pp, 1) if total_pp else 0.0

    # --- Player History --------------------------------------------------
    player_history_js = json.dumps(results_data.get("players", {}))

    # --- Top 25 Results --------------------------------------------------
    t25_dates = sorted(top25_data.get("dates", {}).keys())
    last7_dates = t25_dates[-7:]

    yesterday_cards = []
    daily_hit_rate = None
    daily_date = None
    if t25_dates:
        daily_date = t25_dates[-1]
        yesterday_cards = top25_data["dates"][daily_date]["top25"]
        decided = [e for e in yesterday_cards if e["grade"] in ("win", "loss")]
        if decided:
            hits = sum(1 for e in decided if e["grade"] == "win")
            daily_hit_rate = round(100 * hits / len(decided), 1)
    yesterday_cards_js = json.dumps(yesterday_cards)

    rolling_hits = rolling_decided = 0
    for d in last7_dates:
        for e in top25_data["dates"][d]["top25"]:
            if e["grade"] in ("win", "loss"):
                rolling_decided += 1
                if e["grade"] == "win":
                    rolling_hits += 1
    rolling_hit_rate = round(100 * rolling_hits / rolling_decided, 1) if rolling_decided else None

    record_rows = []
    weekly = []
    trend_symbols = {"win": "✓", "loss": "✗", "push": "~"}
    for pid, p in top25_data.get("players", {}).items():
        history = p.get("history", [])
        appearances = len(history)
        if appearances == 0:
            continue
        hits = sum(1 for h in history if h["grade"] == "win")
        losses = sum(1 for h in history if h["grade"] == "loss")
        decided = hits + losses
        hit_rate = round(100 * hits / decided, 1) if decided else 0.0
        avg_proj = round(sum(h["projected_ud"] for h in history) / appearances, 2)
        avg_actual = round(sum(h["actual_ud"] for h in history) / appearances, 2)
        trend = "".join(trend_symbols[h["grade"]] for h in history[-5:])
        record_rows.append({
            "name": p.get("name", ""),
            "team": p.get("team", ""),
            "times": len(p.get("dates_seen", [])) or appearances,
            "record": f"{hits}-{losses}",
            "hit_rate": hit_rate,
            "avg_proj": avg_proj,
            "avg_actual": avg_actual,
            "trend": trend,
        })

        recent_decided = [h for h in history if h["date"] in last7_dates and h["grade"] in ("win", "loss")]
        if len(recent_decided) >= 2:
            r_hits = sum(1 for h in recent_decided if h["grade"] == "win")
            weekly.append({
                "name": p.get("name", ""),
                "rate": round(100 * r_hits / len(recent_decided), 1),
                "n": len(recent_decided),
            })

    record_rows.sort(key=lambda r: r["hit_rate"], reverse=True)
    record_rows_js = json.dumps(record_rows)

    best_performer = max(weekly, key=lambda x: x["rate"]) if weekly else None
    worst_performer = min(weekly, key=lambda x: x["rate"]) if weekly else None

    # Default display order for every card grid is "soonest game first",
    # which is independent of (and applied after) whatever scoring/edge
    # logic decided which players make a given section - missing game
    # times sort last rather than first.
    def _by_game_time(row_list):
        return sorted(row_list, key=lambda r: r.get("game_date_utc") or "9999")

    # --- Top 25 cards -------------------------------------------------
    cards = [build_card(row) for row in _by_game_time(rows[:25])]
    cards_js = json.dumps(cards)

    # --- Value Plays: model vs. market disagreement of VALUE_PLAY_EDGE+ pts -
    # Only players with a posted UD/PP line count - no line means no market
    # to disagree with. Capped at top 4 OVER + top 4 UNDER by edge size.
    over_rows = [r for r in rows if r.get("market_anchored") and (r.get("edge") or 0) >= VALUE_PLAY_EDGE]
    under_rows = [r for r in rows if r.get("market_anchored") and (r.get("edge") or 0) <= -VALUE_PLAY_EDGE]
    over_rows.sort(key=lambda r: r["edge"], reverse=True)
    under_rows.sort(key=lambda r: r["edge"])
    value_rows = over_rows[:4] + under_rows[:4]
    save_value_plays(date_str, value_rows)
    value_cards = [build_card(row) for row in _by_game_time(value_rows)]
    value_cards_js = json.dumps(value_cards)

    # --- Unanchored: no posted UD/PP line, model-only projection ----------
    unanchored_rows = _by_game_time([r for r in rows[:25] if not r.get("market_anchored")])
    unanchored_cards = [build_card(row) for row in unanchored_rows]
    unanchored_cards_js = json.dumps(unanchored_cards)

    # --- Full leaderboard table ---------------------------------------
    # Color tiers are relative to today's own distribution (top 25% green,
    # next down to the 40th percentile yellow, rest red) rather than fixed
    # point thresholds - ud_pts now lives on the recalibrated/anchored
    # 3.5-9.5 scale, not the old raw model scale those fixed cutoffs were
    # tuned for, so an absolute "8+" cutoff left almost nothing green.
    cols_js = json.dumps([{"key": k, "label": label} for k, label, _ in DASHBOARD_COLS])
    ud_values = [r.get("ud_pts") or 0 for r in rows]
    full_thresholds = (percentile(ud_values, 75), percentile(ud_values, 40))
    table_rows = []
    for rank, row in enumerate(rows, 1):
        cells = [fmt_value(rank if key == "rank" else row[key], kind)
                 for key, _, kind in DASHBOARD_COLS]
        color = "row-" + tier(row.get("ud_pts") or 0, full_thresholds)
        table_rows.append({
            "team": row["team"],
            "cells": cells,
            "color": color,
            "gameDateUtc": row.get("game_date_utc"),
        })
    # Default table order is soonest game first, independent of rank.
    table_rows.sort(key=lambda r: r["gameDateUtc"] or "9999")
    rows_js = json.dumps(table_rows)

    teams = sorted({r["team"] for r in rows})
    teams_js = json.dumps(teams)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>MLB Fantasy Projections — {date_str}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, Segoe UI, Arial, sans-serif; margin: 0; background: #0d1626; color: #e6e9f0; }}

  header {{ background: #0a1628; padding: 18px 24px; display: flex; align-items: center;
            justify-content: space-between; flex-wrap: wrap; gap: 8px;
            border-bottom: 1px solid #1f2c46; }}
  .logo {{ font-size: 22px; font-weight: 800; color: #fff; letter-spacing: 0.5px; }}
  .header-mid {{ font-size: 14px; color: #9fb0cc; text-align: center; }}
  .header-mid .games {{ font-weight: 700; color: #fff; }}
  .header-right {{ font-size: 12px; color: #6c7da0; text-align: right; }}

  .filter-bar {{ padding: 14px 24px 0; display: flex; justify-content: flex-end; }}
  .filter-toggle {{ font-size: 13px; color: #c4cee0; display: flex; align-items: center; gap: 6px; cursor: pointer; }}
  .filter-toggle input {{ width: 15px; height: 15px; cursor: pointer; }}
  .game-time {{ color: #fbbf24; }}

  .tabs {{ padding: 14px 24px 0; display: flex; gap: 8px; }}
  .tab-btn {{ padding: 9px 20px; font-size: 14px; border: 1px solid #2a3a5c; background: #16213a;
              color: #c4cee0; cursor: pointer; border-radius: 8px 8px 0 0; }}
  .tab-btn.active {{ background: #1c2944; color: #fff; border-bottom-color: #1c2944; font-weight: 600; }}

  .panel {{ padding: 18px 24px; }}
  .panel.hidden {{ display: none; }}

  .legend {{ margin-bottom: 14px; font-size: 12px; color: #9fb0cc; }}
  .legend span {{ display: inline-block; padding: 3px 12px; margin-right: 8px; border-radius: 4px; font-weight: 600; }}
  .legend .green  {{ background: #15351f; color: #4ade80; border: 1px solid #4ade80; }}
  .legend .yellow {{ background: #3a3315; color: #fbbf24; border: 1px solid #fbbf24; }}
  .legend .red    {{ background: #3a1818; color: #f87171; border: 1px solid #f87171; }}

  /* Top 25 card grid */
  .card-grid {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 14px; }}
  @media (max-width: 1400px) {{ .card-grid {{ grid-template-columns: repeat(3, 1fr); }} }}
  @media (max-width: 900px)  {{ .card-grid {{ grid-template-columns: repeat(2, 1fr); }} }}
  @media (max-width: 600px)  {{ .card-grid {{ grid-template-columns: 1fr; }} }}

  .card {{ background: #16213a; border-radius: 10px; padding: 14px 16px; border: 3px solid #555;
           transition: transform 0.15s ease, box-shadow 0.15s ease; position: relative; }}
  .card:hover {{ transform: scale(1.04); box-shadow: 0 6px 20px rgba(0,0,0,0.4); z-index: 2; }}
  .card.green  {{ border-color: #4ade80; }}
  .card.yellow {{ border-color: #fbbf24; }}
  .card.red    {{ border-color: #f87171; }}

  .card .name {{ font-size: 18px; font-weight: 800; color: #fff; }}
  .card .meta {{ font-size: 12px; color: #9fb0cc; margin-top: 2px; }}
  .card .pts-row {{ margin-top: 10px; display: flex; gap: 18px; align-items: baseline; }}
  .card .ud-pts {{ font-size: 26px; font-weight: 800; color: #4ade80; }}
  .card .pp-pts {{ font-size: 18px; font-weight: 700; color: #60a5fa; }}
  .card .line-pts {{ font-size: 18px; font-weight: 700; color: #9fb0cc; }}
  .card .pts-label {{ font-size: 10px; color: #6c7da0; display: block; }}
  .t25-call {{ display: inline-flex; align-items: center; gap: 4px; margin-top: 8px; padding: 2px 8px;
               border-radius: 10px; font-weight: 800; font-size: 12px; }}
  .t25-call.over  {{ background: rgba(96,165,250,0.15); color: #60a5fa; }}
  .t25-call.under {{ background: rgba(251,146,60,0.15); color: #fb923c; }}
  .card .stat-line {{ font-size: 12px; color: #c4cee0; margin-top: 8px; }}
  .card .badge {{ display: inline-block; margin-top: 8px; padding: 2px 8px; font-size: 11px;
                  font-weight: 700; color: #0d1626; background: #4ade80; border-radius: 10px; }}
  .card .badge-adjusted {{ background: #60a5fa; margin-left: 6px; }}
  .card .badge-anchored {{ background: #f0abfc; margin-left: 6px; }}
  .card .badge-projected {{ background: #fbbf24; color: #3a2a00; margin-left: 6px; }}
  .card .badge-no-line {{ background: #fb923c; color: #1a0800; margin-left: 6px; }}
  .card .badge-getaway {{ background: #f87171; color: #1a0000; margin-left: 6px; }}

  /* Full leaderboard table */
  .controls {{ display: flex; gap: 10px; margin-bottom: 12px; flex-wrap: wrap; }}
  .controls input, .controls select {{ padding: 8px 12px; font-size: 14px; background: #16213a;
              color: #e6e9f0; border: 1px solid #2a3a5c; border-radius: 6px; }}
  .table-wrap {{ max-height: 75vh; overflow: auto; border: 1px solid #2a3a5c; border-radius: 8px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th, td {{ padding: 8px 12px; font-size: 13px; text-align: left; white-space: nowrap; }}
  th {{ background: #1c2944; color: #fff; cursor: pointer; position: sticky; top: 0; user-select: none;
        border-bottom: 2px solid #2a3a5c; }}
  th:hover {{ background: #25355a; }}
  th.sorted-asc::after {{ content: " \\25B2"; }}
  th.sorted-desc::after {{ content: " \\25BC"; }}
  tbody tr:nth-child(odd)  {{ background: #141d33; }}
  tbody tr:nth-child(even) {{ background: #182241; }}
  tbody tr:hover {{ background: #25355a; }}
  td {{ border-bottom: 1px solid #1f2c46; color: #d6deef; }}

  /* Results tab */
  .results-summary {{ display: flex; gap: 14px; margin-bottom: 16px; }}
  .summary-card {{ background: #16213a; border: 1px solid #2a3a5c; border-radius: 10px;
                   padding: 14px 22px; text-align: center; }}
  .summary-value {{ font-size: 28px; font-weight: 800; color: #fff; }}
  .summary-label {{ font-size: 12px; color: #9fb0cc; margin-top: 4px; }}
  .cal-grid {{ display: flex; flex-wrap: wrap; gap: 8px; }}
  .cal-cell {{ width: 78px; height: 58px; background: #16213a; border: 2px solid #555;
               border-radius: 8px; display: flex; flex-direction: column; align-items: center;
               justify-content: center; }}
  .cal-date {{ font-size: 11px; color: #9fb0cc; }}
  .cal-rate {{ font-size: 16px; font-weight: 800; color: #fff; margin-top: 2px; }}
  .empty-msg {{ color: #9fb0cc; font-size: 14px; }}

  /* Player History tab */
  .ph-player {{ background: #16213a; border: 1px solid #2a3a5c; border-radius: 10px;
                padding: 14px 18px; margin-bottom: 14px; }}
  .ph-player h3 {{ margin: 0 0 10px; color: #fff; font-size: 16px; }}
  .ph-player h3 .meta {{ color: #9fb0cc; font-size: 12px; font-weight: 400; }}
  .ph-table {{ width: 100%; border-collapse: collapse; }}
  .ph-table th, .ph-table td {{ padding: 6px 10px; font-size: 13px; text-align: left; white-space: nowrap; }}
  .ph-table th {{ color: #9fb0cc; border-bottom: 1px solid #2a3a5c; position: static; cursor: default; }}
  .ph-table td {{ color: #d6deef; border-bottom: 1px solid #1f2c46; }}
  .res-win   {{ color: #4ade80; font-weight: 700; }}
  .res-loss  {{ color: #f87171; font-weight: 700; }}
  .res-push  {{ color: #9fb0cc; font-weight: 700; }}
  .res-none  {{ color: #6c7da0; font-weight: 700; }}

  /* Top 25 Results tab */
  .summary-card.best  {{ border-color: #4ade80; }}
  .summary-card.worst {{ border-color: #f87171; }}
  .t25-card.win    {{ border-color: #4ade80; }}
  .t25-card.push   {{ border-color: #fbbf24; }}
  .t25-card.loss   {{ border-color: #f87171; }}
  .t25-card.nodata {{ border-color: #555; }}
  .t25-overlay {{ position: absolute; top: 6px; right: 10px; font-size: 30px; font-weight: 900; }}
  .t25-overlay.win  {{ color: #4ade80; }}
  .t25-overlay.push {{ color: #fbbf24; }}
  .t25-overlay.loss {{ color: #f87171; }}
  .section-title {{ margin: 22px 0 12px; color: #fff; font-size: 16px; }}
  #t25Tbl tbody tr.row-green  td {{ background: #15351f; }}
  #t25Tbl tbody tr.row-yellow td {{ background: #3a3315; }}
  #t25Tbl tbody tr.row-red    td {{ background: #3a1818; }}
  /* Full Leaderboard color rows */
  #tbl tbody tr.row-green  td {{ background: #0d2418; }}
  #tbl tbody tr.row-yellow td {{ background: #2a240e; }}
  #tbl tbody tr.row-red    td {{ background: #2a1010; }}
  #tbl tbody tr.row-green:hover td  {{ background: #15351f; }}
  #tbl tbody tr.row-yellow:hover td {{ background: #3a3315; }}
  #tbl tbody tr.row-red:hover td    {{ background: #3a1818; }}
  .bt-bar-row {{ display: flex; align-items: center; gap: 10px; margin: 6px 0; }}
  .bt-bar-label {{ width: 160px; font-size: 13px; color: #c4cee0; flex-shrink: 0; }}
  .bt-bar {{ flex: 1; height: 14px; background: #16213a; border: 1px solid #2a3a5c; border-radius: 4px; overflow: hidden; }}
  .bt-bar-fill {{ height: 100%; background: #60a5fa; }}
  .bt-bar-value {{ width: 110px; font-size: 12px; color: #9fb0cc; text-align: right; flex-shrink: 0; }}
  .bt-grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin-top: 12px; }}
  @media (max-width: 900px) {{ .bt-grid-2 {{ grid-template-columns: 1fr; }} }}

  /* Edge / Value Plays */
  .card .edge-row {{ margin-top: 8px; font-size: 12px; color: #9fb0cc; }}
  .edge-tag {{ display: inline-flex; align-items: center; gap: 4px; padding: 2px 8px;
               border-radius: 10px; font-weight: 800; font-size: 12px; }}
  .edge-tag.over    {{ background: rgba(74,222,128,0.15); color: #4ade80; }}
  .edge-tag.under   {{ background: rgba(248,113,113,0.15); color: #f87171; }}
  .edge-tag.neutral {{ background: rgba(159,176,204,0.15); color: #9fb0cc; }}

  /* Platoon Matchup */
  .platoon-matchup {{ margin-top: 8px; padding: 8px 10px; border-radius: 8px; font-size: 11px;
                       border-left: 3px solid #9fb0cc; }}
  .platoon-matchup.edge-batter  {{ background: rgba(74,222,128,0.10); border-left-color: #4ade80; }}
  .platoon-matchup.edge-pitcher {{ background: rgba(248,113,113,0.10); border-left-color: #f87171; }}
  .platoon-matchup.edge-neutral {{ background: rgba(159,176,204,0.10); border-left-color: #9fb0cc; }}
  .platoon-title {{ font-weight: 800; margin-bottom: 4px; color: #cbd5e1; }}
  .platoon-row {{ display: flex; flex-direction: column; gap: 2px; color: #9fb0cc; }}

  .value-plays {{ margin: 16px 0 28px; padding: 16px; border: 2px solid #fbbf24;
                   border-radius: 12px; background: linear-gradient(180deg, rgba(251,191,36,0.08), transparent); }}
  .value-plays h2 {{ margin: 0 0 4px; color: #fbbf24; font-size: 20px; }}
  .value-plays .vp-sub {{ color: #c4cee0; font-size: 13px; margin-bottom: 12px; }}
  .value-plays .card {{ border-width: 2px; }}
  .value-plays-empty {{ color: #9fb0cc; font-size: 13px; }}

  .unanchored-section {{ margin: 16px 0 28px; padding: 16px; border: 1px solid #3a4866;
                          border-radius: 12px; }}
  .unanchored-section h2 {{ margin: 0 0 4px; color: #9fb0cc; font-size: 18px; }}
  .unanchored-section .vp-sub {{ color: #7a8aab; font-size: 13px; margin-bottom: 12px; }}
  .unanchored-empty {{ color: #9fb0cc; font-size: 13px; }}

  /* Freshness banner */
  .freshness-banner {{ padding: 10px 24px; font-size: 14px; font-weight: 700; text-align: center; }}
  .freshness-banner.fresh {{ background: #15351f; color: #4ade80; }}
  .freshness-banner.stale {{ background: #3a1818; color: #f87171; }}
  .card .game-date {{ color: #6c7da0; }}
</style>
</head>
<body>
<header>
  <div class="logo">⚾ MLB FANTASY PRO</div>
  <div class="header-mid">{date_str} &mdash; <span class="games">{games_count} games today</span></div>
  <div class="header-right" id="lastUpdated">Last Updated: {last_updated}</div>
</header>

<div id="freshnessBanner" class="freshness-banner"></div>

<div class="filter-bar">
  <label class="filter-toggle">
    <input type="checkbox" id="hideStartedToggle">
    Hide games already started
  </label>
</div>

<div class="value-plays">
  <h2>🎯 Value Plays</h2>
  <div class="vp-sub">Top 4 OVER + top 4 UNDER calls where our model disagrees with the posted UD/PP line by 1.5+ pts &mdash; highest conviction plays.</div>
  <div class="card-grid" id="valueGrid"></div>
</div>

<div class="unanchored-section">
  <h2>Unanchored (no posted line)</h2>
  <div class="vp-sub">Model-only projections &mdash; no UD/PP market line to compare against yet.</div>
  <div class="card-grid" id="unanchoredGrid"></div>
</div>

<div class="tabs">
  <button class="tab-btn active" id="tab-top25" data-tab="top25">Top 25</button>
  <button class="tab-btn" id="tab-full" data-tab="full">Full Leaderboard</button>
  <button class="tab-btn" id="tab-results" data-tab="results">Results</button>
  <button class="tab-btn" id="tab-history" data-tab="history">Player History</button>
  <button class="tab-btn" id="tab-top25results" data-tab="top25results">Top 25 Results</button>
</div>

<div class="panel" id="panel-top25">
  <div class="legend">
    <span class="green">Green border = Elite (8+ pts)</span>
    <span class="yellow">Yellow = Solid (5-8 pts)</span>
    <span class="red">Red = Avoid (under 5 pts)</span>
  </div>
  <div class="card-grid" id="cardGrid"></div>
</div>

<div class="panel hidden" id="panel-full">
  <div class="controls">
    <input id="search" type="text" placeholder="Search player name...">
    <select id="teamFilter">
      <option value="">All Teams</option>
    </select>
  </div>
  <div class="table-wrap">
    <table id="tbl">
      <thead><tr id="hdr"></tr></thead>
      <tbody id="body"></tbody>
    </table>
  </div>
</div>

<div class="panel hidden" id="panel-results">
  <div class="results-summary">
    <div class="summary-card">
      <div class="summary-value">{overall_ud_rate}%</div>
      <div class="summary-label">UD Hit Rate (last {len(result_dates)}d)</div>
    </div>
    <div class="summary-card">
      <div class="summary-value">{overall_pp_rate}%</div>
      <div class="summary-label">PP Hit Rate (last {len(result_dates)}d)</div>
    </div>
  </div>
  <div class="legend">
    <span class="green">Green = 50%+ hit rate</span>
    <span class="yellow">Yellow = 30-50%</span>
    <span class="red">Red = under 30%</span>
  </div>
  <div class="cal-grid" id="calGrid"></div>
</div>

<div class="panel hidden" id="panel-history">
  <div class="controls">
    <input id="phSearch" type="text" placeholder="Search player name...">
  </div>
  <div id="phResults"></div>
</div>

<div class="panel hidden" id="panel-top25results">
  <div class="results-summary" id="t25Summary"></div>
  <div class="legend">
    <span class="green">Green ✓ = won the call (right side of the UD line)</span>
    <span class="red">Red ✗ = lost the call (wrong side of the UD line)</span>
    <span class="yellow">Yellow ~ = push (landed exactly on the line)</span>
    <span style="background:#1c2944;color:#9fb0cc;border:1px solid #555;">Gray = no result yet</span>
  </div>
  <div class="card-grid" id="t25CardGrid"></div>
  <h3 class="section-title">Running Record (all-time Top 25 appearances)</h3>
  <div class="table-wrap">
    <table id="t25Tbl">
      <thead><tr id="t25Hdr"></tr></thead>
      <tbody id="t25Body"></tbody>
    </table>
  </div>
</div>



<script>
const CARDS = {cards_js};
const VALUE_CARDS = {value_cards_js};
const UNANCHORED_CARDS = {unanchored_cards_js};
const COLS = {cols_js};
const ROWS = {rows_js};
const TEAMS = {teams_js};
const CAL = {calendar_js};
const PLAYER_HISTORY = {player_history_js};
const T25_CARDS = {yesterday_cards_js};
const T25_DAILY_DATE = {json.dumps(daily_date)};
const T25_DAILY_RATE = {json.dumps(daily_hit_rate)};
const T25_ROLLING_RATE = {json.dumps(rolling_hit_rate)};
const T25_BEST = {json.dumps(best_performer)};
const T25_WORST = {json.dumps(worst_performer)};
const T25_RECORDS = {record_rows_js};
const GENERATED_AT = {json.dumps(generated_at_iso)};
const GAME_DATE = {json.dumps(date_str)};
const PLAYER_COUNT = {player_count};
const GAMES_COUNT = {games_count};

// --- Freshness indicator ---
function updateFreshness() {{
  const generated = new Date(GENERATED_AT);
  const now = new Date();
  const ageHours = (now - generated) / 3600000;

  const timeStr = generated.toLocaleString('en-US', {{
    month: 'long', day: 'numeric', year: 'numeric', hour: 'numeric', minute: '2-digit', hour12: true,
  }});
  const lastUpdatedEl = document.getElementById('lastUpdated');
  if (ageHours > 4) {{
    lastUpdatedEl.innerHTML = `Last Updated: ${{timeStr}} PT &mdash; <span style="color:#f87171">&#9888; Data may be stale - pipeline may not have run</span>`;
  }} else {{
    lastUpdatedEl.innerHTML = `Last Updated: ${{timeStr}} PT &mdash; <span style="color:#4ade80">&#10003; Fresh</span>`;
  }}

  const banner = document.getElementById('freshnessBanner');
  const todayStr = now.getFullYear() + '-' + String(now.getMonth() + 1).padStart(2, '0') + '-' + String(now.getDate()).padStart(2, '0');
  const updatedTime = generated.toLocaleTimeString('en-US', {{ hour: 'numeric', minute: '2-digit', hour12: true }});
  if (GAME_DATE === todayStr) {{
    banner.className = 'freshness-banner fresh';
    banner.innerHTML = `&#10003; Today's data &mdash; ${{GAME_DATE}} &mdash; ${{PLAYER_COUNT}} players &mdash; ${{GAMES_COUNT}} games &mdash; Updated ${{updatedTime}}`;
  }} else {{
    banner.className = 'freshness-banner stale';
    banner.innerHTML = `&#9888; Showing data from ${{GAME_DATE}} &mdash; pipeline may not have run today`;
  }}
}}
updateFreshness();
setInterval(updateFreshness, 60000);
setInterval(() => location.reload(), 1800000);

// --- Top 25 / Value Play cards ---
function edgeRowHtml(c) {{
  if (c.edgeLabel === null || c.edgeLabel === undefined) return '';
  const sign = c.edge > 0 ? '+' : '';
  if (c.edgeLabel === 'over') {{
    return `<div class="edge-row"><span class="edge-tag over">&#8593; OVER</span> ${{sign}}${{c.edge.toFixed(1)}} vs line ${{c.udLine.toFixed(1)}}</div>`;
  }}
  if (c.edgeLabel === 'under') {{
    return `<div class="edge-row"><span class="edge-tag under">&#8595; UNDER</span> ${{sign}}${{c.edge.toFixed(1)}} vs line ${{c.udLine.toFixed(1)}}</div>`;
  }}
  return `<div class="edge-row"><span class="edge-tag neutral">NEUTRAL</span> ${{sign}}${{c.edge.toFixed(1)}} vs line ${{c.udLine.toFixed(1)}}</div>`;
}}

function platoonMatchupHtml(c) {{
  const pm = c.platoonMatchup;
  if (!pm) return '';
  const cls = pm.advantage === 'batter' ? 'edge-batter' : (pm.advantage === 'pitcher' ? 'edge-pitcher' : 'edge-neutral');
  const arrow = pm.advantage === 'batter' ? '&#9650; Batter edge' : (pm.advantage === 'pitcher' ? '&#9660; Pitcher edge' : '&#9644; Neutral');
  const bWoba = pm.batterWoba != null ? pm.batterWoba.toFixed(3) : 'N/A';
  const pWoba = pm.pitcherWoba != null ? pm.pitcherWoba.toFixed(3) : 'N/A';
  return `
    <div class="platoon-matchup ${{cls}}">
      <div class="platoon-title">Platoon Matchup &middot; ${{arrow}}</div>
      <div class="platoon-row">
        <span>Batter ${{pm.batterLabel || ''}}: <b>${{bWoba}}</b> wOBA</span>
        <span>Pitcher ${{pm.pitcherLabel || ''}}: <b>${{pWoba}}</b> wOBA-against</span>
      </div>
    </div>`;
}}

function renderCard(c) {{
  const card = document.createElement('div');
  card.className = 'card ' + c.tier;
  if (c.gameDateUtc) card.dataset.gameTimeUtc = c.gameDateUtc;
  card.innerHTML = `
    <div class="name">${{c.name}}</div>
    <div class="meta">${{c.team}} &middot; Batting ${{c.order}} &middot; <span class="game-date">${{GAME_DATE}}</span>${{c.gameTimePt ? ` &middot; <span class="game-time">${{c.gameTimePt}}</span>` : ''}}</div>
    <div class="pts-row">
      <div><span class="ud-pts">${{c.ud}}</span><span class="pts-label">UD PTS</span></div>
      <div><span class="pp-pts">${{c.pp}}</span><span class="pts-label">PP PTS</span></div>
    </div>
    <div class="stat-line">xwOBA ${{c.xwoba}} &nbsp;|&nbsp; Barrel% ${{c.barrel}} &nbsp;|&nbsp; Opp ERA ${{c.era}}</div>
    <div class="stat-line">${{c.wxIcon}} ${{c.wxText}} &nbsp;|&nbsp; Park ${{c.park}}</div>
    ${{edgeRowHtml(c)}}
    ${{platoonMatchupHtml(c)}}
    ${{c.platoon ? '<div class="badge">Platoon Edge</div>' : ''}}
    ${{c.adjusted ? '<div class="badge badge-adjusted">Model adjusted</div>' : ''}}
    ${{c.anchored ? '<div class="badge badge-anchored">Live line</div>' : ''}}
    ${{c.projectedLineup && !c.getawayDayRisk ? '<div class="badge badge-projected">&#9888; Projected Lineup</div>' : ''}}
    ${{c.noLinePenalty ? '<div class="badge badge-no-line">&#9888; No Line &ndash; Lower Confidence</div>' : ''}}
    ${{c.getawayDayRisk ? '<div class="badge badge-getaway">&#9888; Projected Lineup &ndash; Getaway Day Risk</div>' : ''}}
  `;
  return card;
}}

const grid = document.getElementById('cardGrid');
for (const c of CARDS) {{
  grid.appendChild(renderCard(c));
}}

// --- Value Plays ---
const valueGrid = document.getElementById('valueGrid');
if (VALUE_CARDS.length === 0) {{
  valueGrid.outerHTML = '<div class="value-plays-empty">No 1.5+ pt disagreements with the market right now.</div>';
}} else {{
  for (const c of VALUE_CARDS) {{
    valueGrid.appendChild(renderCard(c));
  }}
}}

// --- Unanchored (no posted line) ---
const unanchoredGrid = document.getElementById('unanchoredGrid');
if (UNANCHORED_CARDS.length === 0) {{
  unanchoredGrid.outerHTML = '<div class="unanchored-empty">All Top 25 players have a posted UD/PP line.</div>';
}} else {{
  for (const c of UNANCHORED_CARDS) {{
    unanchoredGrid.appendChild(renderCard(c));
  }}
}}

// --- Hide-started-games filter (applies to Top 25, Full Leaderboard,
// Value Plays, Unanchored - NOT Results/Player History, which are
// historical). Single shared toggle so state stays consistent across tabs.
let hideStarted = false;

function gameHasStarted(iso) {{
  if (!iso) return false;
  return new Date(iso).getTime() <= Date.now();
}}

function applyHideStartedFilter() {{
  document.querySelectorAll('[data-game-time-utc]').forEach(el => {{
    el.style.display = (hideStarted && gameHasStarted(el.dataset.gameTimeUtc)) ? 'none' : '';
  }});
}}

const hideStartedToggle = document.getElementById('hideStartedToggle');
hideStartedToggle.addEventListener('change', () => {{
  hideStarted = hideStartedToggle.checked;
  applyHideStartedFilter();
  render(); // re-apply to the Full Leaderboard table too
}});

applyHideStartedFilter();
setInterval(applyHideStartedFilter, 30000); // live re-check as games start, no reload needed

// --- Tabs ---
const PANELS = ['top25', 'full', 'results', 'history', 'top25results'];
document.querySelectorAll('.tab-btn').forEach(btn => {{
  btn.addEventListener('click', () => {{
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    for (const name of PANELS) {{
      document.getElementById('panel-' + name).classList.toggle('hidden', btn.dataset.tab !== name);
    }}
  }});
}});

// --- Results: calendar heatmap ---
function hitColor(rate) {{
  if (rate >= 50) return '#4ade80';
  if (rate >= 30) return '#fbbf24';
  return '#f87171';
}}

const calGrid = document.getElementById('calGrid');
for (const c of CAL) {{
  const cell = document.createElement('div');
  cell.className = 'cal-cell';
  cell.style.borderColor = hitColor(c.ud_hit_rate);
  cell.title = `${{c.date}}\\nUD hit rate: ${{c.ud_hit_rate}}%\\nPP hit rate: ${{c.pp_hit_rate}}%\\nPlayers graded: ${{c.player_count}}`;
  cell.innerHTML = `<div class="cal-date">${{c.date.slice(5)}}</div><div class="cal-rate">${{c.ud_hit_rate}}%</div>`;
  calGrid.appendChild(cell);
}}
if (CAL.length === 0) {{
  calGrid.innerHTML = '<div class="empty-msg">No results yet — run tracker.py after games finish.</div>';
}}

// --- Player History ---
const phSearch = document.getElementById('phSearch');
const phResults = document.getElementById('phResults');

function renderHistory() {{
  const q = phSearch.value.trim().toLowerCase();
  phResults.innerHTML = '';
  if (!q) return;
  const matches = Object.values(PLAYER_HISTORY).filter(p => p.name.toLowerCase().includes(q)).slice(0, 10);
  if (matches.length === 0) {{
    phResults.innerHTML = '<div class="empty-msg">No history found for that player.</div>';
    return;
  }}
  for (const p of matches) {{
    const div = document.createElement('div');
    div.className = 'ph-player';
    const rowsHtml = p.history.slice().reverse().map(h => `
      <tr>
        <td>${{h.date}}${{h.game_time_pt ? ` <span class="meta">${{h.game_time_pt}}</span>` : ''}}</td>
        <td>${{h.projected_ud != null ? h.projected_ud.toFixed(1) : 'N/A'}}</td>
        <td>${{h.ud_line != null ? h.ud_line.toFixed(1) : '—'}}</td>
        <td>${{h.actual_ud}}</td>
        <td class="res-${{h.result_ud || 'none'}}">${{h.result_ud || '—'}}</td>
        <td>${{h.projected_pp != null ? h.projected_pp.toFixed(1) : 'N/A'}}</td>
        <td>${{h.pp_line != null ? h.pp_line.toFixed(1) : '—'}}</td>
        <td>${{h.actual_pp}}</td>
        <td class="res-${{h.result_pp || 'none'}}">${{h.result_pp || '—'}}</td>
      </tr>`).join('');
    div.innerHTML = `
      <h3>${{p.name}} <span class="meta">${{p.team}}</span></h3>
      <table class="ph-table">
        <thead><tr><th>Date</th><th>Proj UD</th><th>UD Line</th><th>Actual UD</th><th>UD Result</th><th>Proj PP</th><th>PP Line</th><th>Actual PP</th><th>PP Result</th></tr></thead>
        <tbody>${{rowsHtml}}</tbody>
      </table>`;
    phResults.appendChild(div);
  }}
}}

phSearch.addEventListener('input', renderHistory);

// --- Top 25 Results ---
const t25Summary = document.getElementById('t25Summary');
{{
  let html = '';
  if (T25_DAILY_DATE) {{
    const decidedCount = T25_CARDS.filter(c => c.grade === 'win' || c.grade === 'loss').length;
    const hitCount = T25_CARDS.filter(c => c.grade === 'win').length;
    html += `<div class="summary-card"><div class="summary-value">${{hitCount}}/${{decidedCount}}</div>
             <div class="summary-label">${{T25_DAILY_DATE}}: wins (${{T25_DAILY_RATE !== null ? T25_DAILY_RATE : 'N/A'}}%)</div></div>`;
  }} else {{
    html += `<div class="summary-card"><div class="summary-value">N/A</div>
             <div class="summary-label">No Top 25 results yet</div></div>`;
  }}
  html += `<div class="summary-card"><div class="summary-value">${{T25_ROLLING_RATE !== null ? T25_ROLLING_RATE + '%' : 'N/A'}}</div>
           <div class="summary-label">Last 7 days hit rate</div></div>`;
  if (T25_BEST) {{
    html += `<div class="summary-card best"><div class="summary-value">${{T25_BEST.rate}}%</div>
             <div class="summary-label">Best this week: ${{T25_BEST.name}}</div></div>`;
  }}
  if (T25_WORST) {{
    html += `<div class="summary-card worst"><div class="summary-value">${{T25_WORST.rate}}%</div>
             <div class="summary-label">Worst this week: ${{T25_WORST.name}}</div></div>`;
  }}
  t25Summary.innerHTML = html;
}}

const t25Grid = document.getElementById('t25CardGrid');
for (const c of T25_CARDS) {{
  const card = document.createElement('div');
  let borderClass = 'nodata';
  let overlay = '';
  if (c.grade === 'win') {{
    borderClass = 'win';
    overlay = '<div class="t25-overlay win">&#10003;</div>';
  }} else if (c.grade === 'push') {{
    borderClass = 'push';
    overlay = '<div class="t25-overlay push">~</div>';
  }} else if (c.grade === 'loss') {{
    borderClass = 'loss';
    overlay = '<div class="t25-overlay loss">&#10007;</div>';
  }}
  let callBadge = '';
  if (c.ud_line != null && c.projected_ud != null) {{
    if (c.projected_ud > c.ud_line) {{
      callBadge = '<div class="t25-call over">&#8593; OVER</div>';
    }} else if (c.projected_ud < c.ud_line) {{
      callBadge = '<div class="t25-call under">&#8595; UNDER</div>';
    }}
  }}
  card.className = 'card t25-card ' + borderClass;
  card.innerHTML = `
    ${{overlay}}
    <div class="name">${{c.name}}</div>
    <div class="meta">${{c.team}} &middot; Batting ${{c.order}} &middot; <span class="game-date">${{c.date}}</span>${{c.gameTimePt ? ` &middot; <span class="game-time">${{c.gameTimePt}}</span>` : ''}}</div>
    <div class="pts-row">
      <div><span class="ud-pts">${{c.ud}}</span><span class="pts-label">PROJ UD</span></div>
      <div><span class="line-pts">${{c.ud_line != null ? c.ud_line.toFixed(1) : 'N/A'}}</span><span class="pts-label">UD LINE</span></div>
      <div><span class="pp-pts">${{c.actual_ud !== null ? c.actual_ud : 'N/A'}}</span><span class="pts-label">ACTUAL UD</span></div>
    </div>
    ${{callBadge}}
    <div class="stat-line">xwOBA ${{c.xwoba}} &nbsp;|&nbsp; Barrel% ${{c.barrel}} &nbsp;|&nbsp; Opp ERA ${{c.era}}</div>
    <div class="stat-line">${{c.wxIcon}} ${{c.wxText}} &nbsp;|&nbsp; Park ${{c.park}}</div>
    ${{c.platoon ? '<div class="badge">Platoon Edge</div>' : ''}}
    ${{c.adjusted ? '<div class="badge badge-adjusted">Model adjusted</div>' : ''}}
    ${{c.noLinePenalty ? '<div class="badge badge-no-line">&#9888; No Line &ndash; Lower Confidence</div>' : ''}}
    ${{c.getawayDayRisk ? '<div class="badge badge-getaway">&#9888; Projected Lineup &ndash; Getaway Day Risk</div>' : ''}}
  `;
  t25Grid.appendChild(card);
}}
if (T25_CARDS.length === 0) {{
  t25Grid.innerHTML = '<div class="empty-msg">No Top 25 results yet — run tracker.py after games finish.</div>';
}}

const T25_COLS = [
  {{key: 'name',       label: 'Player'}},
  {{key: 'times',      label: 'Times in Top 25'}},
  {{key: 'record',     label: 'Record'}},
  {{key: 'hit_rate',   label: 'Hit Rate %'}},
  {{key: 'avg_proj',   label: 'Avg Projected'}},
  {{key: 'avg_actual', label: 'Avg Actual'}},
  {{key: 'trend',      label: 'Trend (last 5)'}},
];
const t25Hdr = document.getElementById('t25Hdr');
T25_COLS.forEach((c, i) => {{
  const th = document.createElement('th');
  th.textContent = c.label;
  th.addEventListener('click', () => sortT25(i));
  t25Hdr.appendChild(th);
}});

let t25SortCol = 3; // Hit Rate %
let t25SortDir = -1;

function sortT25(i) {{
  if (t25SortCol === i) {{
    t25SortDir *= -1;
  }} else {{
    t25SortCol = i;
    t25SortDir = -1;
  }}
  renderT25();
}}

function recordRowClass(rate) {{
  if (rate >= 60) return 'row-green';
  if (rate >= 40) return 'row-yellow';
  return 'row-red';
}}

function renderT25() {{
  const key = T25_COLS[t25SortCol].key;
  const sorted = T25_RECORDS.slice().sort((a, b) => {{
    const av = a[key], bv = b[key];
    let cmp;
    if (typeof av === 'number' && typeof bv === 'number') {{
      cmp = av - bv;
    }} else {{
      cmp = String(av).localeCompare(String(bv));
    }}
    return cmp * t25SortDir;
  }});

  Array.from(t25Hdr.children).forEach((th, i) => {{
    th.classList.remove('sorted-asc', 'sorted-desc');
    if (i === t25SortCol) th.classList.add(t25SortDir === 1 ? 'sorted-asc' : 'sorted-desc');
  }});

  const body = document.getElementById('t25Body');
  body.innerHTML = '';
  for (const r of sorted) {{
    const tr = document.createElement('tr');
    tr.className = recordRowClass(r.hit_rate);
    for (const c of T25_COLS) {{
      const td = document.createElement('td');
      td.textContent = c.key === 'hit_rate' ? r[c.key] + '%' : r[c.key];
      tr.appendChild(td);
    }}
    body.appendChild(tr);
  }}
  if (sorted.length === 0) {{
    body.innerHTML = '<tr><td class="empty-msg">No graded Top 25 history yet.</td></tr>';
  }}
}}

renderT25();

// --- Full leaderboard table ---
const hdr = document.getElementById('hdr');
COLS.forEach((c, i) => {{
  const th = document.createElement('th');
  th.textContent = c.label;
  th.dataset.idx = i;
  th.addEventListener('click', () => sortBy(i));
  hdr.appendChild(th);
}});

const teamFilter = document.getElementById('teamFilter');
for (const t of TEAMS) {{
  const opt = document.createElement('option');
  opt.value = t;
  opt.textContent = t;
  teamFilter.appendChild(opt);
}}

const GAME_TIME_COL_IDX = COLS.findIndex(c => c.key === 'game_time_pt');
let sortCol = GAME_TIME_COL_IDX >= 0 ? GAME_TIME_COL_IDX : 4;
let sortDir = 1; // ascending = soonest game first, by default

function sortBy(i) {{
  if (sortCol === i) {{
    sortDir *= -1;
  }} else {{
    sortCol = i;
    sortDir = i === GAME_TIME_COL_IDX ? 1 : -1;
  }}
  render();
}}

function render() {{
  const filter = document.getElementById('search').value.trim().toLowerCase();
  const team = teamFilter.value;
  let rows = ROWS.filter(r => String(r.cells[1]).toLowerCase().includes(filter));
  if (team) rows = rows.filter(r => r.team === team);
  if (hideStarted) rows = rows.filter(r => !gameHasStarted(r.gameDateUtc));

  rows = rows.slice().sort((a, b) => {{
    if (sortCol === GAME_TIME_COL_IDX) {{
      return ((a.gameDateUtc || '9999').localeCompare(b.gameDateUtc || '9999')) * sortDir;
    }}
    const av = a.cells[sortCol], bv = b.cells[sortCol];
    const an = parseFloat(av), bn = parseFloat(bv);
    let cmp;
    if (!isNaN(an) && !isNaN(bn)) {{
      cmp = an - bn;
    }} else {{
      cmp = String(av).localeCompare(String(bv));
    }}
    return cmp * sortDir;
  }});

  Array.from(hdr.children).forEach((th, i) => {{
    th.classList.remove('sorted-asc', 'sorted-desc');
    if (i === sortCol) th.classList.add(sortDir === 1 ? 'sorted-asc' : 'sorted-desc');
  }});

  const body = document.getElementById('body');
  body.innerHTML = '';
  for (const r of rows) {{
    const tr = document.createElement('tr');
    if (r.color) tr.classList.add(r.color);
    for (const c of r.cells) {{
      const td = document.createElement('td');
      td.textContent = c;
      tr.appendChild(td);
    }}
    body.appendChild(tr);
  }}
}}

document.getElementById('search').addEventListener('input', render);
teamFilter.addEventListener('change', render);
render();
</script>
</body>
</html>
"""
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    return out_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def prepare_dashboard_context(date_arg=None):
    """Load and process the latest projection data, returning everything
    needed to render the dashboard: rows, the date they're for, and the
    results/top25 history used by the Results / Player History tabs."""
    players, date_str = proj.load_data(date_arg)

    results_path = os.path.join("data", "results", "all_results.json")
    results_data = None
    if os.path.exists(results_path):
        with open(results_path, encoding="utf-8") as f:
            results_data = json.load(f)

    rows = [build_row(p) for p in players]
    recalibrate_points(rows)

    market_lines = get_market_lines(date_str)
    market_corrections = (results_data or {}).get("market_corrections", {})
    anchored_ids = apply_market_anchor(rows, market_lines, market_corrections)

    corrections = build_corrections(results_data)
    apply_corrections(rows, corrections, skip=anchored_ids)
    apply_no_line_penalty(rows, anchored_ids)
    rows.sort(key=lambda r: r["ud_pts"], reverse=True)

    top25_path = os.path.join("data", "results", "top25_results.json")
    top25_data = None
    if os.path.exists(top25_path):
        with open(top25_path, encoding="utf-8") as f:
            top25_data = json.load(f)

    return rows, date_str, results_data, top25_data


def regenerate_dashboard(date_arg=None):
    """Rebuild output/dashboard.html from the latest projection + results
    data and push it to GitHub Pages. Used by tracker.py after nightly
    grading so the Results and Player History tabs update without a full
    pipeline run."""
    rows, date_str, results_data, top25_data = prepare_dashboard_context(date_arg)
    html_path = os.path.join("output", "dashboard.html")
    write_dashboard(rows, date_str, html_path, results_data, top25_data)
    print(f"Dashboard saved -> {os.path.abspath(html_path)}")
    deploy_to_github_pages(html_path, date_str)
    return html_path


def main():
    date_arg = sys.argv[1] if len(sys.argv) > 1 else None
    rows, date_str, results_data, top25_data = prepare_dashboard_context(date_arg)
    players, _ = proj.load_data(date_arg)

    ud_values = [r["ud_pts"] for r in rows]
    ud_thresholds = (percentile(ud_values, 75), percentile(ud_values, 40))

    wb = Workbook()
    wb.remove(wb.active)

    # Tab 1: Top Plays
    write_leaderboard_sheet(wb, "Top Plays", rows[:25], ud_thresholds, color_rows=True)

    # Tab 2: Full Leaderboard
    write_leaderboard_sheet(wb, "Full Leaderboard", rows, ud_thresholds, color_rows=True, autofilter=True)

    # Tab 3: Matchup Report
    players_by_game = defaultdict(list)
    for r in rows:
        players_by_game[r["game_pk"]].append(r)
    write_matchup_sheet(wb, players_by_game)

    # Tab 4: SP Report
    write_sp_sheet(wb, players)

    # Tab 5: Stack Targets
    players_by_team = defaultdict(list)
    for r in rows:
        players_by_team[r["team"]].append(r)
    write_stack_sheet(wb, players_by_team)

    os.makedirs("output", exist_ok=True)
    xlsx_path = os.path.join("output", f"MLB_Projections_{date_str}.xlsx")
    wb.save(xlsx_path)
    print(f"Excel report saved -> {os.path.abspath(xlsx_path)}")

    html_path = os.path.join("output", "dashboard.html")
    write_dashboard(rows, date_str, html_path, results_data, top25_data)
    print(f"Dashboard saved -> {os.path.abspath(html_path)}")

    deploy_to_github_pages(html_path, date_str)

    # Auto-open (skipped for unattended/scheduled runs)
    if not os.environ.get("MLB_HEADLESS"):
        os.startfile(os.path.abspath(xlsx_path))
        webbrowser.open(f"file:///{os.path.abspath(html_path)}")


# ---------------------------------------------------------------------------
# GitHub Pages auto-deploy
#
# Copies the freshly generated dashboard to docs/index.html and pushes it to
# the mlb-fantasy repo so https://macassvic-cmd.github.io/mlb-fantasy/ stays
# in sync with the latest run. Best-effort: any failure (no network, no git,
# merge conflicts, etc.) is logged and swallowed so it never breaks the
# pipeline.
# ---------------------------------------------------------------------------

DOCS_DASHBOARD_PATH = os.path.join("docs", "index.html")


GIT_SUBPROCESS_TIMEOUT = 60  # seconds - git push has hung indefinitely under
# the S4U scheduled-task context (no interactive desktop to satisfy a
# credential prompt), blocking the whole tracker.py/pipeline.py process for
# hours with nothing to time it out. Every git call here is now bounded.


def deploy_to_github_pages(html_path, date_str):
    try:
        os.makedirs("docs", exist_ok=True)
        shutil.copyfile(html_path, DOCS_DASHBOARD_PATH)

        repo_root = os.path.dirname(os.path.abspath(__file__))
        subprocess.run(["git", "add", "docs/index.html"], cwd=repo_root, check=True,
                        capture_output=True, text=True, timeout=GIT_SUBPROCESS_TIMEOUT)

        status = subprocess.run(["git", "status", "--porcelain", "docs/index.html"],
                                 cwd=repo_root, check=True, capture_output=True, text=True,
                                 timeout=GIT_SUBPROCESS_TIMEOUT)
        if not status.stdout.strip():
            print("GitHub Pages: no dashboard changes to deploy.")
            return

        subprocess.run(["git", "commit", "-q", "-m", f"Update dashboard for {date_str}"],
                        cwd=repo_root, check=True, capture_output=True, text=True,
                        timeout=GIT_SUBPROCESS_TIMEOUT)
        subprocess.run(["git", "push"], cwd=repo_root, check=True, capture_output=True, text=True,
                        timeout=GIT_SUBPROCESS_TIMEOUT)
        print("GitHub Pages: dashboard deployed -> https://macassvic-cmd.github.io/mlb-fantasy/")
    except Exception as e:
        if isinstance(e, subprocess.TimeoutExpired):
            detail = f"timed out after {GIT_SUBPROCESS_TIMEOUT}s"
        elif isinstance(e, subprocess.CalledProcessError):
            detail = e.stderr
        else:
            detail = str(e)
        print(f"GitHub Pages deploy skipped (non-fatal): {detail}")


if __name__ == "__main__":
    main()
